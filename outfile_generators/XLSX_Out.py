import openpyxl

class XLSX_Out:
    def __init__(self, data, filepath):
        outfile = openpyxl.Workbook()
        sheet = outfile.active

        for i in range(len(data[0])):
            sheet.cell(row=1, column=i+1).value = data[0][i]
            sheet.cell(row=1, column=i+1).font = openpyxl.styles.Font(bold=True)

        for row in data[1:]:
            sheet.append(row)

        outfile.save(filepath)