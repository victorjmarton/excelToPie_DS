import openpyxl

class Client_List:
    def __init__(self, filepath, last_stages):
        self.filepath = filepath
        self.names = []
        self.last_stages = last_stages
        self.client_list = self.load_client_list()

    # Loads and formats the Client List file while storing sheet names
    def load_client_list(self):
        client_list = openpyxl.load_workbook(self.filepath)
        self.names = client_list.sheetnames
        return self.reformat(client_list)

    # Converts data from the .xlsx file into a workable list
    def reformat(self, worksheet : openpyxl.workbook.workbook.Workbook):
        sheets = []

        for sheet in worksheet.sheetnames:
            sheets.append(list(worksheet[sheet].values))
        return sheets